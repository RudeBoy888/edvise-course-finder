#!/usr/bin/env python3
"""
Scrape course descriptions from top 20 provider websites by matching course names
"""

import json
import openpyxl
from pathlib import Path
from collections import defaultdict
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, quote
import time

# Load institutions and courses from Excel
def get_top_providers_with_courses(limit=20):
    """Get top N providers by number of courses"""
    excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel not found")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Load institutions
        institutions = {}
        ws_inst = wb['Institutions']
        for row_idx in range(4, ws_inst.max_row + 1):
            provider_code = ws_inst.cell(row=row_idx, column=1).value
            name = ws_inst.cell(row=row_idx, column=2).value or ws_inst.cell(row=row_idx, column=3).value
            website = ws_inst.cell(row=row_idx, column=6).value

            if provider_code and name:
                institutions[str(provider_code).strip()] = {
                    "name": str(name).strip(),
                    "website": str(website).strip() if website else ""
                }

        # Load courses and count by provider
        ws_courses = wb['Courses']
        provider_course_count = defaultdict(list)

        for row_idx in range(4, ws_courses.max_row + 1):
            expired = ws_courses.cell(row=row_idx, column=24).value
            if str(expired).strip().lower() != "no":
                continue

            provider_code = ws_courses.cell(row=row_idx, column=1).value
            course_name = ws_courses.cell(row=row_idx, column=4).value

            if provider_code and course_name:
                provider_code = str(provider_code).strip()
                if provider_code in institutions:
                    provider_course_count[provider_code].append(str(course_name).strip())

        # Get top 20 by course count
        top_providers = sorted(
            provider_course_count.items(),
            key=lambda x: len(x[1]),
            reverse=True
        )[:limit]

        result = []
        for provider_code, courses in top_providers:
            inst = institutions.get(provider_code, {})
            result.append({
                "provider_code": provider_code,
                "name": inst.get("name", "Unknown"),
                "website": inst.get("website", ""),
                "course_count": len(courses),
                "courses": courses[:5]  # First 5 courses as sample
            })

        return result

    except Exception as e:
        print(f"❌ Error: {e}")
        return []

def search_course_on_website(provider_name, provider_website, course_name):
    """Try to find course description on provider's website"""
    try:
        if not provider_website or not provider_website.startswith('http'):
            return None

        # Try common course page patterns
        search_patterns = [
            f"{provider_website}/courses/{quote(course_name)}",
            f"{provider_website}/course/{quote(course_name)}",
            f"{provider_website}/programs/{quote(course_name)}",
            f"{provider_website}/programs/",  # List page
        ]

        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }

        # Try search on the main website
        search_url = f"{provider_website}?s={quote(course_name)}"

        response = requests.get(search_url, headers=headers, timeout=10)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Try to extract text about the course
            text_content = soup.get_text()

            # Look for paragraphs containing course name
            for p in soup.find_all(['p', 'div']):
                p_text = p.get_text()
                if course_name.lower() in p_text.lower():
                    cleaned = ' '.join(p_text.split())
                    if len(cleaned) > 50:  # Only if substantial
                        return cleaned[:300]  # First 300 chars

            return None

        return None

    except requests.exceptions.Timeout:
        return "timeout"
    except Exception as e:
        return f"error: {str(e)}"

def main():
    print("🔍 Scraping Top 20 Providers for Course Descriptions\n")
    print("=" * 70)

    # Get top providers
    print("\n📊 Finding top 20 providers by course count...\n")
    providers = get_top_providers_with_courses(limit=20)

    if not providers:
        print("❌ No providers found!")
        return

    print(f"✅ Found {len(providers)} top providers:\n")

    results = []

    for idx, provider in enumerate(providers, 1):
        name = provider['name'][:50]
        website = provider['website'][:50] if provider['website'] else "N/A"
        count = provider['course_count']

        print(f"[{idx}/20] {name:<50} | {count:>3} courses")

        # Try to scrape a sample course from this provider
        if provider['courses'] and provider['website']:
            sample_course = provider['courses'][0]
            print(f"         Searching for: '{sample_course[:40]}'...", end=" ", flush=True)

            desc = search_course_on_website(
                provider['name'],
                provider['website'],
                sample_course
            )

            if desc and desc not in ["timeout", None] and not desc.startswith("error"):
                print(f"✅ Found!")
                results.append({
                    "provider": provider['name'],
                    "website": provider['website'],
                    "course": sample_course,
                    "description": desc
                })
            elif desc == "timeout":
                print(f"⏱️  Timeout")
            else:
                print(f"❌ Not found")

        time.sleep(1)  # Be respectful

    # Save results
    print("\n" + "=" * 70)
    print(f"\n📊 RESULTS:\n")
    print(f"✅ Successfully found: {len(results)}/20 providers")

    if results:
        print("\n📝 SAMPLE DESCRIPTIONS FOUND:\n")
        for item in results[:5]:
            print(f"Provider: {item['provider']}")
            print(f"Course: {item['course']}")
            print(f"Desc: {item['description'][:100]}...")
            print()

    # Save to JSON
    output_file = Path("/Users/rudybobek/edvise-course-finder/provider_descriptions_found.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            "summary": {
                "providers_tested": len(providers),
                "descriptions_found": len(results),
                "success_rate": f"{len(results)/len(providers)*100:.1f}%"
            },
            "results": results,
            "all_providers": providers
        }, f, indent=2, ensure_ascii=False)

    print(f"💾 Results saved to: {output_file}")

    # Assessment
    print("\n" + "=" * 70)
    print("\n💡 ASSESSMENT:\n")

    success_rate = len(results) / len(providers) * 100
    if success_rate >= 40:
        print(f"✅ Success rate: {success_rate:.0f}% - This approach works!")
        print("   Next step: Manually review and add to Excel")
    elif success_rate >= 20:
        print(f"⚠️  Partial success: {success_rate:.0f}%")
        print("   Many provider websites are too dynamic")
        print("   Recommendation: Mix of scraping + manual")
    else:
        print(f"❌ Low success: {success_rate:.0f}%")
        print("   Most provider websites require JavaScript")
        print("   Recommendation: Manual entry for now")

if __name__ == "__main__":
    main()
