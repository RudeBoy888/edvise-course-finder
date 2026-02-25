#!/usr/bin/env python3
"""
Scrape training.gov.au descriptions by matching course NAMES (not codes)
Uses Selenium for JavaScript rendering
Only accepts confident matches
"""

import openpyxl
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import json
from difflib import SequenceMatcher

def get_top_courses_by_provider(limit_per_provider=3, max_providers=20):
    """Get top courses from each provider to test"""
    excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel not found")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        institutions = {}
        ws_inst = wb['Institutions']
        for row_idx in range(4, ws_inst.max_row + 1):
            provider_code = ws_inst.cell(row=row_idx, column=1).value
            name = ws_inst.cell(row=row_idx, column=2).value or ws_inst.cell(row=row_idx, column=3).value
            if provider_code:
                institutions[str(provider_code).strip()] = str(name).strip() if name else "Unknown"

        ws_courses = wb['Courses']
        courses_by_provider = {}

        for row_idx in range(4, ws_courses.max_row + 1):
            expired = ws_courses.cell(row=row_idx, column=24).value
            if str(expired).strip().lower() != "no":
                continue

            provider_code = str(ws_courses.cell(row=row_idx, column=1).value).strip()
            course_name = ws_courses.cell(row=row_idx, column=4).value

            if provider_code and course_name:
                if provider_code not in courses_by_provider:
                    courses_by_provider[provider_code] = []

                if len(courses_by_provider[provider_code]) < limit_per_provider:
                    courses_by_provider[provider_code].append({
                        "provider_name": institutions.get(provider_code, "Unknown"),
                        "course_name": str(course_name).strip()
                    })

        # Flatten and limit
        all_courses = []
        for provider_code, courses in list(courses_by_provider.items())[:max_providers]:
            all_courses.extend(courses)

        return all_courses

    except Exception as e:
        print(f"❌ Error: {e}")
        return []

def search_course_on_nrt(course_name, driver):
    """Search for course on training.gov.au and extract description"""
    try:
        # Go to search page
        search_url = f"https://training.gov.au/search?searchText={course_name}&searchType=NRT"
        driver.get(search_url)

        # Wait for search results
        time.sleep(2)

        # Get page source
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # Look for course links
        course_links = soup.find_all('a', href=True)

        matches = []
        for link in course_links:
            href = link.get('href', '')
            text = link.get_text(strip=True)

            if '/training/details/' in href and text:
                # Calculate similarity
                similarity = SequenceMatcher(None, course_name.lower(), text.lower()).ratio()

                if similarity > 0.7:  # 70% match confidence
                    matches.append({
                        "name": text,
                        "url": href if href.startswith('http') else f"https://training.gov.au{href}",
                        "similarity": similarity
                    })

        if not matches:
            return None

        # Take best match
        best_match = max(matches, key=lambda x: x['similarity'])

        if best_match['similarity'] < 0.85:  # Only very confident matches
            return None

        # Fetch description from course page
        driver.get(best_match['url'])
        time.sleep(1)

        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # Look for qualification description
        description = None
        for heading in soup.find_all(['h2', 'h3', 'h4']):
            if 'Qualification description' in heading.get_text():
                next_elem = heading.find_next(['p', 'div'])
                if next_elem:
                    description = next_elem.get_text(strip=True)[:500]
                    break

        # Fallback: look for any substantial paragraph
        if not description:
            for p in soup.find_all('p'):
                text = p.get_text(strip=True)
                if len(text) > 100 and 'qualification' in text.lower():
                    description = text[:500]
                    break

        if description:
            return {
                "found_name": best_match['name'],
                "url": best_match['url'],
                "similarity": best_match['similarity'],
                "description": description
            }

        return None

    except Exception as e:
        return None

def main():
    print("🔍 Scraping training.gov.au by course NAME (Selenium)\n")
    print("=" * 70)

    # Get test courses
    print("\n📚 Getting sample courses...\n")
    courses = get_top_courses_by_provider(limit_per_provider=2, max_providers=10)

    if not courses:
        print("❌ No courses found!")
        return

    print(f"✅ Found {len(courses)} sample courses\n")

    # Setup Selenium
    print("🌐 Starting Selenium browser...\n")
    chrome_options = Options()
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')

    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
    except Exception as e:
        print(f"❌ Could not start Selenium: {e}")
        print("   Install Chrome browser or use: brew install chromium")
        return

    # Search for courses
    results = {
        "found": [],
        "not_found": [],
        "errors": []
    }

    for idx, course in enumerate(courses, 1):
        name = course["course_name"]
        print(f"[{idx}/{len(courses)}] Searching: '{name[:50]}...'", end=" ", flush=True)

        result = search_course_on_nrt(name, driver)

        if result:
            results["found"].append({
                "course_name": name,
                "provider": course["provider_name"],
                **result
            })
            print(f"✅ FOUND! ({result['similarity']:.0%})")
        else:
            results["not_found"].append(name)
            print(f"❌ Not found or low confidence")

        time.sleep(1)

    driver.quit()

    # Results
    print("\n" + "=" * 70)
    print(f"\n📊 RESULTS:\n")
    print(f"✅ Found: {len(results['found'])}/{len(courses)}")
    print(f"❌ Not found: {len(results['not_found'])}/{len(courses)}")

    if results["found"]:
        print(f"\n✅ MATCHES:\n")
        for item in results["found"]:
            print(f"  Course: {item['course_name'][:60]}")
            print(f"  Found: {item['found_name']}")
            print(f"  Match: {item['similarity']:.0%}")
            print(f"  Desc: {item['description'][:80]}...")
            print()

    # Save results
    output_file = Path("/Users/rudybobek/edvise-course-finder/nrt_selenium_results.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"💾 Saved to: {output_file}")

    # Assessment
    print("\n" + "=" * 70)
    success_rate = len(results["found"]) / len(courses) * 100
    print(f"\n💡 ASSESSMENT: {success_rate:.0f}% success rate\n")

    if success_rate >= 50:
        print("✅ Method works! Can scale to more courses")
        print("   Next: Run for all unique course names")
    elif success_rate >= 20:
        print("⚠️  Partial success - some courses match")
        print("   Names may differ slightly between systems")
    else:
        print("❌ Low success - courses not matching by name")

if __name__ == "__main__":
    main()
