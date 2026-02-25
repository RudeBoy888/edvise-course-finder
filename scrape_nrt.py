#!/usr/bin/env python3
"""
Scrape NRT qualification descriptions from training.gov.au
Tests matching between CRICOS codes and NRT codes
"""

import json
import time
import openpyxl
from pathlib import Path
import requests
from bs4 import BeautifulSoup

# First, extract course codes from Excel to test
def get_course_codes_from_excel(limit=50):
    """Extract first N course codes from Excel"""
    excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel not found at {excel_path}")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Courses']

        codes = set()
        for row_idx in range(4, min(4 + limit * 2, ws.max_row + 1)):  # Get more rows to account for duplicates
            code = ws.cell(row=row_idx, column=3).value  # courseCode is column 3
            name = ws.cell(row=row_idx, column=4).value
            expired = ws.cell(row=row_idx, column=24).value

            # Only active courses
            if code and str(expired).strip().lower() == "no" and len(codes) < limit:
                codes.add(str(code).strip())

        return sorted(list(codes))
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return []

def fetch_nrt_page(code):
    """Fetch training.gov.au page for a specific code"""
    url = f"https://training.gov.au/training/details/{code}/qualdetails"

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }

        response = requests.get(url, headers=headers, timeout=10)

        if response.status_code == 404:
            return {"status": "not_found", "code": code}

        if response.status_code == 200:
            return {"status": "ok", "code": code, "content": response.text}

        return {"status": f"error_{response.status_code}", "code": code}

    except requests.exceptions.Timeout:
        return {"status": "timeout", "code": code}
    except requests.exceptions.RequestException as e:
        return {"status": f"error", "code": code, "error": str(e)}

def extract_description(html, code):
    """Try to extract qualification description from HTML"""
    try:
        soup = BeautifulSoup(html, 'html.parser')

        # Try to find description in various ways
        # Look for text containing "Qualification description"

        # Method 1: Look for heading + following text
        for h in soup.find_all(['h2', 'h3', 'h4']):
            if 'Qualification description' in h.get_text():
                # Get next paragraph or content
                next_elem = h.find_next(['p', 'div'])
                if next_elem:
                    return next_elem.get_text(strip=True)[:500]  # First 500 chars

        # Method 2: Look in any div/section with description text
        for elem in soup.find_all(['div', 'section']):
            text = elem.get_text()
            if 'trade qualification' in text.lower() or 'This is a' in text:
                return text.strip()[:500]

        return None
    except Exception as e:
        print(f"  Error parsing HTML: {e}")
        return None

def main():
    print("🔍 Testing NRT Code Scraping\n")
    print("=" * 60)

    # Step 1: Get codes from Excel
    print("\n📊 Extracting course codes from Excel...")
    codes = get_course_codes_from_excel(limit=50)

    if not codes:
        print("❌ No codes found!")
        return

    print(f"✅ Found {len(codes)} unique course codes")
    print(f"   Sample: {codes[:5]}\n")

    # Step 2: Test fetching
    print("=" * 60)
    print("\n🌐 Testing NRT lookups...\n")

    results = {
        "found": [],
        "not_found": [],
        "errors": []
    }

    for idx, code in enumerate(codes[:10], 1):  # Test first 10
        print(f"[{idx}/10] Testing {code}...", end=" ")

        response = fetch_nrt_page(code)

        if response["status"] == "ok":
            desc = extract_description(response["content"], code)
            if desc:
                results["found"].append({
                    "code": code,
                    "description": desc
                })
                print(f"✅ Found (description: {len(desc)} chars)")
            else:
                results["found"].append({
                    "code": code,
                    "description": "[HTML found but no description extracted]"
                })
                print(f"⚠️  Found page but couldn't extract description")

        elif response["status"] == "not_found":
            results["not_found"].append(code)
            print(f"❌ Not found on training.gov.au")

        else:
            results["errors"].append({"code": code, "status": response["status"]})
            print(f"⚠️  Error: {response['status']}")

        time.sleep(1)  # Be respectful to the server

    # Step 3: Print results
    print("\n" + "=" * 60)
    print("\n📊 TEST RESULTS:\n")

    print(f"✅ Successfully found: {len(results['found'])}/10")
    print(f"❌ Not found on NRT: {len(results['not_found'])}/10")
    print(f"⚠️  Errors: {len(results['errors'])}/10")

    if results["found"]:
        print("\n✅ FOUND DESCRIPTIONS:")
        for item in results["found"]:
            print(f"\n  Code: {item['code']}")
            print(f"  Description: {item['description'][:100]}...")

    if results["not_found"]:
        print(f"\n❌ NOT FOUND (codes don't match NRT):")
        print(f"  {results['not_found']}")

    # Save results
    output_file = Path("/Users/rudybobek/edvise-course-finder/nrt_test_results.json")
    with open(output_file, 'w') as f:
        json.dump({
            "summary": {
                "total_tested": 10,
                "found": len(results["found"]),
                "not_found": len(results["not_found"]),
                "errors": len(results["errors"]),
                "success_rate": f"{len(results['found']) / 10 * 100:.1f}%"
            },
            "results": results
        }, f, indent=2)

    print(f"\n💾 Results saved to: {output_file}")

    # Recommendation
    print("\n" + "=" * 60)
    print("\n💡 RECOMMENDATION:\n")

    success_rate = len(results["found"]) / 10 * 100
    if success_rate > 50:
        print(f"✅ Success rate is {success_rate:.0f}% - Scraping is viable!")
        print("   Can proceed with full dataset scraping.")
    else:
        print(f"❌ Success rate is only {success_rate:.0f}% - Most codes don't match NRT")
        print("   CRICOS and NRT codes are different systems.")
        print("   Would need name-based matching instead.")

if __name__ == "__main__":
    main()
