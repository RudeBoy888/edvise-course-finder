#!/usr/bin/env python3
"""
Generate realistic descriptions for ALL courses based on course metadata
Uses course name, level, field, and duration to create professional descriptions
"""

import openpyxl
from pathlib import Path
import csv
import re

def get_all_courses_without_descriptions():
    """Get ALL courses without descriptions from Excel"""
    excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel not found")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Load institutions
        institutions = {}
        ws_inst = wb['Institutions']
        for row_idx in range(4, ws_inst.max_row + 1):
            provider_code = ws_inst.cell(row=row_idx, column=1).value
            name = ws_inst.cell(row=row_idx, column=2).value or ws_inst.cell(row=row_idx, column=3).value

            if provider_code:
                institutions[str(provider_code).strip()] = str(name).strip() if name else "Unknown"

        # Load courses without descriptions
        ws_courses = wb['Courses']
        courses = []

        for row_idx in range(4, ws_courses.max_row + 1):
            expired = ws_courses.cell(row=row_idx, column=24).value
            if str(expired).strip().lower() != "no":
                continue

            provider_code = ws_courses.cell(row=row_idx, column=1).value
            course_name = ws_courses.cell(row=row_idx, column=4).value
            course_level = ws_courses.cell(row=row_idx, column=13).value
            field = ws_courses.cell(row=row_idx, column=7).value
            duration = ws_courses.cell(row=row_idx, column=20).value
            description = ws_courses.cell(row=row_idx, column=25).value

            # Only show courses WITHOUT descriptions
            if provider_code and course_name and (not description or str(description).strip() == ""):
                provider_code = str(provider_code).strip()

                # Extract field name (strip "NN - " prefix)
                field_clean = ""
                if field:
                    field_str = str(field).strip()
                    if " - " in field_str:
                        field_clean = field_str.split(" - ", 1)[1]
                    else:
                        field_clean = field_str

                courses.append({
                    "row": row_idx,
                    "provider_code": provider_code,
                    "provider_name": institutions.get(provider_code, "Unknown"),
                    "course_name": str(course_name).strip(),
                    "course_level": str(course_level).strip() if course_level else "",
                    "field": field_clean,
                    "duration": int(duration) if duration and str(duration).isdigit() else None,
                })

        return courses

    except Exception as e:
        print(f"❌ Error: {e}")
        return []

def generate_description(course):
    """Generate a realistic description based on course metadata"""
    name = course["course_name"]
    level = course["course_level"].lower()
    field = course["field"].lower()
    duration = course["duration"]

    # Determine description based on course level
    if "bachelor" in level:
        desc = f"Develop comprehensive expertise in {field or 'your chosen field'}. This degree program covers advanced theory and practical application. Ideal for graduates seeking professional careers in {field or 'related industries'}."

    elif "master" in level or "postgraduate" in level:
        desc = f"Advance your professional expertise with specialized knowledge in {field or 'your discipline'}. Develop research and critical thinking skills. Prepare for leadership roles in {field or 'related sectors'}."

    elif "graduate diploma" in level:
        desc = f"Gain specialized qualifications in {field or 'your field of interest'}. Develop practical and theoretical knowledge. Suitable for professionals advancing their careers or those transitioning into {field or 'the industry'}."

    elif "diploma" in level:
        desc = f"Develop practical expertise in {field or 'your chosen area'}. Learn industry-standard practices and technical skills. Ideal for career advancement and professional development in {field or 'related fields'}."

    elif "certificate" in level and "iv" in level.lower():
        desc = f"Gain professional qualification in {field or 'your chosen trade or profession'}. Develop advanced skills and knowledge. Prepare for supervisory and specialized roles in {field or 'the industry'}."

    elif "certificate" in level and "iii" in level.lower():
        desc = f"Develop practical trade or technical skills in {field or 'your field of interest'}. Learn hands-on techniques and industry practices. Foundation for career progression in {field or 'related professions'}."

    elif "certificate" in level and "ii" in level.lower():
        desc = f"Build fundamental skills and knowledge in {field or 'your chosen field'}. Develop basic competencies for entry-level positions. Pathway to further qualifications in {field or 'related areas'}."

    elif "certificate" in level or "award" in level.lower():
        desc = f"Develop foundational competency in {field or 'your area of study'}. Learn essential skills and knowledge. Suitable for entry-level positions and further study in {field or 'related fields'}."

    elif "foundation" in level.lower():
        desc = f"Prepare for university study with foundational knowledge in {field or 'your chosen discipline'}. Build academic skills and understanding. Pathway to bachelor degree programs."

    elif "english" in name.lower() or "esl" in name.lower():
        level_map = {
            "elementary": "Build foundational English communication skills for everyday use and basic study.",
            "pre-intermediate": "Develop practical English language abilities for work and academic environments.",
            "intermediate": "Enhance intermediate English proficiency for study and professional communication.",
            "upper-intermediate": "Master advanced English for academic and professional success.",
            "advanced": "Achieve proficiency in advanced English for university and professional contexts.",
            "academic": "Specialize in academic English required for university study.",
            "ielts": "Prepare comprehensively for IELTS examination with targeted strategies.",
            "pearson": "Prepare for Pearson PTE Academic examination with intensive practice."
        }

        for key, val in level_map.items():
            if key in name.lower():
                return val

        desc = "Develop English language skills for academic and professional advancement."

    else:
        desc = f"Develop competency and expertise in {field or 'your chosen field'}. Learn practical and theoretical knowledge. Suitable for career development in {field or 'related industries'}."

    return desc.strip()

def main():
    print("🔄 Generating descriptions for ALL courses\n")
    print("=" * 70)

    # Get all courses
    print("\n🔍 Finding all courses without descriptions...\n")
    courses = get_all_courses_without_descriptions()

    if not courses:
        print("✅ All courses already have descriptions!")
        return

    print(f"✅ Found {len(courses)} courses without descriptions\n")
    print(f"📝 Generating descriptions... (this may take a moment)\n")

    # Generate descriptions
    descriptions_generated = []
    for idx, course in enumerate(courses, 1):
        desc = generate_description(course)
        descriptions_generated.append({
            **course,
            "description": desc
        })

        if idx % 100 == 0:
            print(f"   ✅ Generated {idx}/{len(courses)} descriptions...")

    print(f"\n✅ Generated {len(descriptions_generated)} descriptions")

    # Export to CSV
    print("\n📝 Creating CSV template...")
    output_file = Path("/Users/rudybobek/edvise-course-finder/courses_all_descriptions.csv")

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("Row,Provider,Course Code,Course Name,Level,Field,Duration (weeks),Description\n")

        for course in descriptions_generated:
            provider = course["provider_name"].replace(",", ";")
            course_name = course["course_name"].replace(",", ";")
            level = course["course_level"]
            field = course["field"]
            duration = course["duration"] or ""
            desc = course["description"]

            f.write(f'{course["row"]},"{provider}","","{course_name}","{level}","{field}",{duration},"{desc}"\n')

    print(f"   ✅ {output_file}")

    # Summary
    print("\n" + "=" * 70)
    print(f"\n✅ READY TO IMPORT:\n")
    print(f"   Generated: {len(descriptions_generated)} descriptions")
    print(f"   CSV file: {output_file}\n")
    print(f"🔄 NEXT STEPS:\n")
    print(f"   1. python3 import_all_descriptions.py")
    print(f"   2. python3 convert_excel_to_json.py")
    print(f"   3. Refresh app\n")

if __name__ == "__main__":
    main()
