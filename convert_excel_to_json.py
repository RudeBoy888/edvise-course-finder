#!/usr/bin/env python3
"""
Convert CRICOS Excel data to JSON format for the course finder app.
"""

import json
import openpyxl
from pathlib import Path
from urllib.parse import urlparse
import time

def extract_domain(website_url):
    """Extract domain from full URL.
    Examples:
        https://www.unsw.edu.au → unsw.edu.au
        https://example.com → example.com
        "" → None
    """
    if not website_url or not str(website_url).strip():
        return None

    try:
        parsed = urlparse(str(website_url).strip())
        domain = parsed.netloc or parsed.path

        # Remove 'www.' prefix if present
        if domain.startswith('www.'):
            domain = domain[4:]

        return domain if domain else None
    except:
        return None

def load_logo_mapping():
    """Load local logo mapping from JSON file."""
    import json
    try:
        with open('src/assets/logo-mapping.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def get_logo_url(domain, logo_mapping):
    """Get logo URL for a domain.

    First checks local mapping (src/assets/logos/).
    Falls back to empty string for unmapped institutions (will use placeholder).

    Returns logo URL or empty string.
    """
    if not domain:
        return ""

    # Check if domain has local logo
    if domain in logo_mapping:
        return logo_mapping[domain]

    # No mapping - will use placeholder
    return ""

def load_institutions(ws, logo_mapping):
    """Load institutions from Institutions sheet.
    Returns dict: { providerCode: { name, website, domain, logoUrl } }
    """
    institutions = {}
    for row_idx in range(4, ws.max_row + 1):
        provider_code = ws.cell(row=row_idx, column=1).value
        trading_name = ws.cell(row=row_idx, column=2).value
        institution_name = ws.cell(row=row_idx, column=3).value
        website = ws.cell(row=row_idx, column=6).value

        if not provider_code:
            continue

        # Use trading name if available, otherwise institution name
        name = trading_name.strip() if trading_name and trading_name.strip() else institution_name
        registered_name = institution_name.strip() if institution_name and institution_name.strip() else None

        # Extract domain and get logo URL (from mapping or empty)
        domain = extract_domain(website)
        logo_url = get_logo_url(domain, logo_mapping)

        institution_data = {
            "name": name,
            "website": website if website else "",
            "domain": domain if domain else "",
            "logoUrl": logo_url if logo_url else ""
        }

        # Add registered name if different from trading name
        if registered_name and registered_name != name:
            institution_data["registeredName"] = registered_name

        institutions[str(provider_code).strip()] = institution_data

    return institutions

def load_locations(ws):
    """Load locations from Locations sheet.
    Returns dict: { (providerCode, locationName): { address, city, state, postcode } }
    """
    locations = {}
    for row_idx in range(4, ws.max_row + 1):
        provider_code = ws.cell(row=row_idx, column=1).value
        location_name = ws.cell(row=row_idx, column=3).value
        address_line1 = ws.cell(row=row_idx, column=5).value
        address_line2 = ws.cell(row=row_idx, column=6).value
        address_line3 = ws.cell(row=row_idx, column=7).value
        city = ws.cell(row=row_idx, column=9).value
        state = ws.cell(row=row_idx, column=10).value
        postcode = ws.cell(row=row_idx, column=11).value

        if not provider_code or not location_name:
            continue

        # Build full address
        address_parts = [address_line1]
        if address_line2:
            address_parts.append(address_line2)
        if address_line3:
            address_parts.append(address_line3)
        address = ", ".join(p.strip() for p in address_parts if p)

        key = (str(provider_code).strip(), str(location_name).strip())
        locations[key] = {
            "address": address,
            "city": city if city else "",
            "state": state if state else "",
            "postcode": str(postcode) if postcode else ""
        }

    return locations

def load_courses(ws):
    """Load courses from Courses sheet.
    Filters out expired courses (Expired == "No").
    Returns list of { providerCode, courseCode, courseName, durationWeeks, tuitionFee, nonTuitionFee, totalCost, courseLevel, fieldOfEducation, isFoundationStudies, hasWorkComponent, description }
    """
    courses = []
    for row_idx in range(4, ws.max_row + 1):
        expired = ws.cell(row=row_idx, column=24).value

        # Skip expired courses
        if str(expired).strip().lower() != "no":
            continue

        provider_code = ws.cell(row=row_idx, column=1).value
        course_code = ws.cell(row=row_idx, column=3).value
        course_name = ws.cell(row=row_idx, column=4).value
        duration_weeks = ws.cell(row=row_idx, column=20).value
        tuition_fee = ws.cell(row=row_idx, column=21).value
        non_tuition_fee = ws.cell(row=row_idx, column=22).value
        total_cost = ws.cell(row=row_idx, column=23).value
        field_of_education = ws.cell(row=row_idx, column=7).value
        course_level = ws.cell(row=row_idx, column=13).value
        foundation_studies = ws.cell(row=row_idx, column=14).value
        work_component = ws.cell(row=row_idx, column=15).value
        description = ws.cell(row=row_idx, column=25).value  # New: Description column

        if not provider_code or not course_code:
            continue

        # Extract fieldOfEducation: strip "NN - " prefix if present
        field_str = ""
        if field_of_education:
            field_str = str(field_of_education).strip()
            if field_str and " - " in field_str:
                # Remove numeric prefix (e.g., "01 - Natural and Physical Sciences" → "Natural and Physical Sciences")
                parts = field_str.split(" - ", 1)
                if len(parts) == 2:
                    field_str = parts[1]

        # Convert boolean fields
        is_foundation = str(foundation_studies).strip().lower() in ["yes", "1", "true"] if foundation_studies else False
        has_work = str(work_component).strip().lower() in ["yes", "1", "true"] if work_component else False

        courses.append({
            "providerCode": str(provider_code).strip(),
            "courseCode": str(course_code).strip(),
            "courseName": course_name if course_name else "",
            "durationWeeks": int(duration_weeks) if duration_weeks and str(duration_weeks).isdigit() else None,
            "tuitionFee": int(tuition_fee) if tuition_fee and isinstance(tuition_fee, (int, float)) else None,
            "nonTuitionFee": int(non_tuition_fee) if non_tuition_fee and isinstance(non_tuition_fee, (int, float)) else None,
            "totalCost": int(total_cost) if total_cost and isinstance(total_cost, (int, float)) else None,
            "courseLevel": str(course_level).strip() if course_level else "",
            "fieldOfEducation": field_str,
            "isFoundationStudies": is_foundation,
            "hasWorkComponent": has_work,
            "description": str(description).strip() if description else ""
        })

    return courses

def load_course_locations(ws):
    """Load course locations from Course Locations sheet.
    Groups by (providerCode, courseCode).
    Returns dict: { (providerCode, courseCode): [{ locationName, city, state }, ...] }
    """
    course_locations = {}
    for row_idx in range(4, ws.max_row + 1):
        provider_code = ws.cell(row=row_idx, column=1).value
        course_code = ws.cell(row=row_idx, column=3).value
        location_name = ws.cell(row=row_idx, column=4).value
        city = ws.cell(row=row_idx, column=5).value
        state = ws.cell(row=row_idx, column=6).value

        if not provider_code or not course_code or not location_name:
            continue

        key = (str(provider_code).strip(), str(course_code).strip())
        if key not in course_locations:
            course_locations[key] = []

        # Avoid duplicates
        location_entry = {
            "locationName": str(location_name).strip(),
            "city": city if city else "",
            "state": state if state else ""
        }

        # Only add if not already present
        if location_entry not in course_locations[key]:
            course_locations[key].append(location_entry)

    return course_locations

def main():
    excel_path = Path("/Users/rudybobek/Desktop/cricos_new.xlsx")
    output_path = Path("/Users/rudybobek/edvise-course-finder/public/courses_data.json")

    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        return

    print("Loading logo mapping...")
    logo_mapping = load_logo_mapping()
    print(f"  Found {len(logo_mapping)} local logos")

    print("Loading workbook...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=False)
    except PermissionError:
        print("  Trying alternative method...")
        import subprocess
        subprocess.run(["chmod", "644", str(excel_path)], check=False)
        wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=False)

    print("Loading institutions...")
    institutions = load_institutions(wb['Institutions'], logo_mapping)
    print(f"  Loaded {len(institutions)} institutions")

    print("Loading locations...")
    locations = load_locations(wb['Locations'])
    print(f"  Loaded {len(locations)} locations")

    print("Loading courses...")
    courses_list = load_courses(wb['Courses'])
    print(f"  Loaded {len(courses_list)} active courses")

    print("Loading course locations...")
    course_locations = load_course_locations(wb['Course Locations'])
    print(f"  Loaded {len(course_locations)} course-location mappings")

    print("Building output structure...")

    # Group courses by provider
    courses_by_provider = {}
    for course in courses_list:
        provider_code = course["providerCode"]
        if provider_code not in courses_by_provider:
            courses_by_provider[provider_code] = []
        courses_by_provider[provider_code].append(course)

    # Build final output
    output = []

    for provider_code, institution_data in institutions.items():
        if provider_code not in courses_by_provider:
            continue  # Skip institutions with no active courses

        # Get all courses for this provider
        provider_courses = courses_by_provider[provider_code]

        # Build course objects with locations
        course_objects = []
        all_states_set = set()

        for course in provider_courses:
            course_code = course["courseCode"]
            location_key = (provider_code, course_code)

            # Get locations for this course
            course_locs = course_locations.get(location_key, [])

            # Group locations by state
            locations_by_state = {}
            for loc in course_locs:
                state = loc["state"]
                if state not in locations_by_state:
                    locations_by_state[state] = []

                # Get full location data
                full_loc_key = (provider_code, loc["locationName"])
                if full_loc_key in locations:
                    loc_data = locations[full_loc_key]
                    location_obj = {
                        "locationName": loc["locationName"],
                        "address": loc_data["address"],
                        "city": loc["city"],
                        "state": state,
                        "postcode": loc_data["postcode"]
                    }
                else:
                    # Fallback if not found in locations table
                    location_obj = {
                        "locationName": loc["locationName"],
                        "address": "",
                        "city": loc["city"],
                        "state": state,
                        "postcode": ""
                    }

                locations_by_state[state].append(location_obj)
                all_states_set.add(state)

            # Only add course if it has locations
            if locations_by_state:
                course_obj = {
                    "courseCode": course["courseCode"],
                    "courseName": course["courseName"],
                    "durationWeeks": course["durationWeeks"],
                    "tuitionFee": course["tuitionFee"],
                    "nonTuitionFee": course["nonTuitionFee"],
                    "totalCost": course["totalCost"],
                    "courseLevel": course["courseLevel"],
                    "fieldOfEducation": course["fieldOfEducation"],
                    "isFoundationStudies": course["isFoundationStudies"],
                    "hasWorkComponent": course["hasWorkComponent"],
                    "description": course["description"],
                    "locations": locations_by_state
                }
                course_objects.append(course_obj)

        # Deduplicate courses by: courseName + price + duration + level
        # Keep first occurrence, remove duplicates
        seen_courses = set()
        deduplicated_courses = []

        for course_obj in course_objects:
            # Create unique key based on identifying fields
            course_key = (
                course_obj["courseName"],
                course_obj["tuitionFee"],
                course_obj["totalCost"],
                course_obj["durationWeeks"],
                course_obj["courseLevel"]
            )

            if course_key not in seen_courses:
                seen_courses.add(course_key)
                deduplicated_courses.append(course_obj)

        # Only add institution if it has courses with locations
        if deduplicated_courses:
            institution_obj = {
                "providerCode": provider_code,
                "name": institution_data["name"],
                "website": institution_data["website"],
                "domain": institution_data["domain"],
                "logoUrl": institution_data["logoUrl"],
                "allStates": sorted(list(all_states_set)),
                "courses": deduplicated_courses
            }
            output.append(institution_obj)

    print(f"Built output with {len(output)} institutions")

    # Create output directory if needed
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Writing to {output_path}...")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"✓ Successfully generated {output_path}")

    # Print some stats
    total_courses = sum(len(inst["courses"]) for inst in output)
    duplicates_removed = len(courses_list) - total_courses
    print(f"\nSummary:")
    print(f"  Institutions: {len(output)}")
    print(f"  Total courses: {total_courses}")
    print(f"  Duplicates removed: {duplicates_removed}")
    print(f"  File size: {output_path.stat().st_size / 1024:.1f} KB")

if __name__ == "__main__":
    main()
