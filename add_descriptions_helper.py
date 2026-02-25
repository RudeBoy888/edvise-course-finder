#!/usr/bin/env python3
"""
Interactive helper to add course descriptions to Excel
Shows top 30 courses without descriptions
"""

import openpyxl
from pathlib import Path
import json
from collections import defaultdict

def get_courses_without_descriptions(limit=30):
    """Get courses without descriptions from Excel"""
    excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel not found")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Load institutions
        institutions = {}
        ws_inst = wb['Institutions']
        for row_idx in range(4, ws_inst.max_row + 1):
            provider_code = ws_inst.cell(row=row_idx, column=1).value
            name = ws_inst.cell(row=row_idx, column=2).value or ws_inst.cell(row=row_idx, column=3).value
            website = ws_inst.cell(row=row_idx, column=6).value

            if provider_code:
                institutions[str(provider_code).strip()] = {
                    "name": str(name).strip() if name else "Unknown",
                    "website": str(website).strip() if website else ""
                }

        # Load courses without descriptions
        ws_courses = wb['Courses']
        courses = []

        for row_idx in range(4, ws_courses.max_row + 1):
            expired = ws_courses.cell(row=row_idx, column=24).value
            if str(expired).strip().lower() != "no":
                continue

            provider_code = ws_courses.cell(row=row_idx, column=1).value
            course_name = ws_courses.cell(row=row_idx, column=4).value
            course_code = ws_courses.cell(row=row_idx, column=3).value
            description = ws_courses.cell(row=row_idx, column=25).value
            course_level = ws_courses.cell(row=row_idx, column=13).value

            # Only show courses WITHOUT descriptions
            if provider_code and course_name and (not description or str(description).strip() == ""):
                provider_code = str(provider_code).strip()
                if provider_code in institutions:
                    courses.append({
                        "row": row_idx,
                        "provider_code": provider_code,
                        "provider_name": institutions[provider_code]["name"],
                        "provider_website": institutions[provider_code]["website"],
                        "course_code": str(course_code).strip() if course_code else "",
                        "course_name": str(course_name).strip(),
                        "course_level": str(course_level).strip() if course_level else "",
                    })

            if len(courses) >= limit:
                break

        return courses

    except Exception as e:
        print(f"❌ Error: {e}")
        return []

def export_to_csv(courses):
    """Export courses to CSV for easy editing"""
    output_file = Path("/Users/rudybobek/edvise-course-finder/courses_to_add_descriptions.csv")

    with open(output_file, 'w', encoding='utf-8') as f:
        # Header
        f.write("Row,Provider,Website,Course Code,Course Name,Level,Description\n")

        for course in courses:
            website = course["provider_website"]
            provider = course["provider_name"].replace(",", ";")
            course_code = course["course_code"]
            course_name = course["course_name"].replace(",", ";")
            level = course["course_level"]

            f.write(f'{course["row"]},"{provider}","{website}","{course_code}","{course_name}","{level}","[PASTE DESCRIPTION HERE]"\n')

    return output_file

def create_import_guide(courses):
    """Create a guide for manual entry"""
    guide_file = Path("/Users/rudybobek/edvise-course-finder/DESCRIPTIONS_GUIDE.md")

    with open(guide_file, 'w', encoding='utf-8') as f:
        f.write("# How to Add Course Descriptions\n\n")
        f.write("## Quick Start\n\n")
        f.write("1. Open: `courses_to_add_descriptions.csv`\n")
        f.write("2. For each course:\n")
        f.write("   - Visit the provider website (click the link)\n")
        f.write("   - Find the course page\n")
        f.write("   - Copy the course description/overview\n")
        f.write("   - Paste it in the CSV (Description column)\n")
        f.write("3. When done, run: `python3 import_descriptions.py`\n\n")

        f.write("## Top 30 Courses Without Descriptions\n\n")
        f.write("| # | Provider | Website | Course Name | Level |\n")
        f.write("|---|----------|---------|-------------|-------|\n")

        for idx, course in enumerate(courses[:30], 1):
            provider = course["provider_name"][:40]
            website = course["provider_website"][:50] if course["provider_website"] else "N/A"
            course_name = course["course_name"][:50]
            level = course["course_level"][:20]

            f.write(f"| {idx} | {provider} | [{website}]({website}) | {course_name} | {level} |\n")

        f.write("\n\n## Tips\n\n")
        f.write("- Look for sections: 'Course Overview', 'Course Description', 'About this course'\n")
        f.write("- Keep descriptions 100-300 characters\n")
        f.write("- Use provider's official language\n")
        f.write("- Don't include pricing/fees in description\n")

    return guide_file

def main():
    print("📚 Course Descriptions Helper\n")
    print("=" * 70)

    # Get courses
    print("\n🔍 Finding courses without descriptions...\n")
    courses = get_courses_without_descriptions(limit=30)

    if not courses:
        print("✅ All courses already have descriptions!")
        return

    print(f"✅ Found {len(courses)} courses without descriptions\n")

    # Export to CSV
    print("📝 Creating CSV template...")
    csv_file = export_to_csv(courses)
    print(f"   ✅ {csv_file}")

    # Create guide
    print("📖 Creating guide...")
    guide_file = create_import_guide(courses)
    print(f"   ✅ {guide_file}")

    # Show sample
    print("\n" + "=" * 70)
    print("\n📋 SAMPLE COURSES (First 5):\n")

    for idx, course in enumerate(courses[:5], 1):
        print(f"{idx}. {course['course_name']}")
        print(f"   Provider: {course['provider_name']}")
        print(f"   Website: {course['provider_website']}")
        print(f"   Level: {course['course_level']}")
        print()

    # Instructions
    print("=" * 70)
    print("\n✅ FILES CREATED:\n")
    print(f"   📊 {csv_file}")
    print(f"      ↳ Edit this file with descriptions\n")
    print(f"   📖 {guide_file}")
    print(f"      ↳ Instructions and list of courses\n")

    print("\n🔄 NEXT STEPS:\n")
    print("   1. Open: courses_to_add_descriptions.csv")
    print("   2. Visit each provider website (link in CSV)")
    print("   3. Copy course description")
    print("   4. Paste in CSV (Description column)")
    print("   5. Save CSV")
    print("   6. Run: python3 import_descriptions.py\n")

if __name__ == "__main__":
    main()
