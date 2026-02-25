#!/usr/bin/env python3
"""
Scrape NRT qualification descriptions by course NAME (not code)
Search training.gov.au for matching course names
"""

import json
import time
import openpyxl
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote

# Extract course names from Excel
def get_courses_from_excel(limit=50):
    """Extract course names and codes from Excel"""
    excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel not found at {excel_path}")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Courses']

        courses = []
        seen = set()

        for row_idx in range(4, ws.max_row + 1):
            if len(courses) >= limit:
                break

            code = ws.cell(row=row_idx, column=3).value  # courseCode
            name = ws.cell(row=row_idx, column=4).value  # courseName
            expired = ws.cell(row=row_idx, column=24).value

            # Only active courses
            if code and name and str(expired).strip().lower() == "no":
                name_clean = str(name).strip()
                if name_clean not in seen:
                    courses.append({
                        "code": str(code).strip(),
                        "name": name_clean
                    })
                    seen.add(name_clean)

        return courses
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return []

def search_nrt_by_name(course_name):
    """Search training.gov.au for a course by name"""
    try:
        # Try to search via the search endpoint
        search_url = f"https://training.gov.au/search"

        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
        }

        params = {
            'searchText': course_name,
            'searchType': 'NRT'
        }

        response = requests.get(search_url, params=params, headers=headers, timeout=15)

        if response.status_code == 200:
            return {
                "status": "ok",
                "name": course_name,
                "url": response.url,
                "content": response.text
            }
        else:
            return {
                "status": f"error_{response.status_code}",
                "name": course_name
            }

    except requests.exceptions.Timeout:
        return {"status": "timeout", "name": course_name}
    except Exception as e:
        return {"status": "error", "name": course_name, "error": str(e)}

def extract_first_result(html, course_name):
    """Extract first search result and its details"""
    try:
        soup = BeautifulSoup(html, 'html.parser')

        # Look for search results
        # Try various selectors for course result links
        result_link = None

        # Method 1: Look for course title links
        for link in soup.find_all('a', href=True):
            href = link.get('href')
            text = link.get_text(strip=True)

            # Check if link points to training.gov.au/training/details/
            if '/training/details/' in href and href.startswith('/'):
                result_link = href
                result_title = text
                break

        if not result_link:
            return None

        # Now fetch the details page
        full_url = f"https://training.gov.au{result_link}"
        response = requests.get(full_url, timeout=15)

        if response.status_code == 200:
            detail_soup = BeautifulSoup(response.text, 'html.parser')

            # Try to extract description from details page
            description = None

            # Look for text content that describes the qualification
            for text in detail_soup.stripped_strings:
                if 'qualification' in text.lower() or 'this is' in text.lower():
                    description = text
                    break

            return {
                "match": result_title,
                "url": full_url,
                "description": description[:300] if description else None
            }

        return None

    except Exception as e:
        print(f"    Error parsing: {e}")
        return None

def main():
    print("🔍 Testing NRT Search by Course NAME\n")
    print("=" * 70)

    # Step 1: Get courses from Excel
    print("\n📚 Extracting course names from Excel...")
    courses = get_courses_from_excel(limit=50)

    if not courses:
        print("❌ No courses found!")
        return

    print(f"✅ Found {len(courses)} unique courses")
    print(f"   Sample: {courses[:3]}\n")

    # Step 2: Test searching
    print("=" * 70)
    print("\n🌐 Testing NRT searches by name...\n")

    results = {
        "matched": [],
        "no_results": [],
        "errors": []
    }

    for idx, course in enumerate(courses[:15], 1):  # Test first 15
        name = course["name"]
        print(f"[{idx}/15] Searching: '{name[:50]}...' ", end="", flush=True)

        search_result = search_nrt_by_name(name)

        if search_result["status"] == "ok":
            # Try to extract first result
            match = extract_first_result(search_result["content"], name)

            if match:
                results["matched"].append({
                    "cricos_name": name,
                    "nrt_match": match["match"],
                    "nrt_url": match["url"],
                    "description": match["description"]
                })
                print(f"✅ MATCHED!")
            else:
                results["no_results"].append(name)
                print(f"⚠️  No match found")

        elif search_result["status"] == "timeout":
            results["errors"].append({"name": name, "error": "timeout"})
            print(f"⏱️  Timeout")

        else:
            results["errors"].append({"name": name, "error": search_result["status"]})
            print(f"❌ Error: {search_result['status']}")

        time.sleep(2)  # Be respectful to server

    # Step 3: Print results
    print("\n" + "=" * 70)
    print("\n📊 TEST RESULTS:\n")

    matched_count = len(results["matched"])
    no_results_count = len(results["no_results"])
    errors_count = len(results["errors"])
    total = 15

    print(f"✅ Matched courses: {matched_count}/15 ({matched_count/total*100:.0f}%)")
    print(f"⚠️  No results found: {no_results_count}/15 ({no_results_count/total*100:.0f}%)")
    print(f"❌ Errors/Timeouts: {errors_count}/15 ({errors_count/total*100:.0f}%)")

    if results["matched"]:
        print("\n✅ SUCCESSFUL MATCHES:\n")
        for item in results["matched"][:5]:
            print(f"  CRICOS: {item['cricos_name'][:60]}")
            print(f"  NRT:    {item['nrt_match']}")
            if item["description"]:
                print(f"  Desc:   {item['description'][:80]}...")
            print()

    if results["no_results"]:
        print(f"\n⚠️  NO RESULTS (courses not found on NRT):")
        for name in results["no_results"][:5]:
            print(f"  - {name[:70]}")

    # Save results
    output_file = Path("/Users/rudybobek/edvise-course-finder/nrt_test_by_name_results.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            "summary": {
                "total_tested": 15,
                "matched": matched_count,
                "no_results": no_results_count,
                "errors": errors_count,
                "success_rate": f"{matched_count / total * 100:.1f}%"
            },
            "matched_courses": results["matched"],
            "no_results": results["no_results"],
            "errors": results["errors"]
        }, f, indent=2, ensure_ascii=False)

    print(f"\n💾 Results saved to: {output_file}")

    # Recommendation
    print("\n" + "=" * 70)
    print("\n💡 ASSESSMENT:\n")

    success_rate = matched_count / total * 100
    if success_rate >= 60:
        print(f"✅ SUCCESS RATE: {success_rate:.0f}% - Search by name WORKS!")
        print("   Recommendation: Proceed with full scraping")
        print("   Next step: Scrape all 25k courses and extract descriptions")

    elif success_rate >= 30:
        print(f"⚠️  PARTIAL SUCCESS: {success_rate:.0f}% - Some courses match")
        print("   Recommendation: Hybrid approach - manual + scraping")
        print("   Next step: Scrape matching courses, add rest manually")

    else:
        print(f"❌ LOW SUCCESS: {success_rate:.0f}% - Most courses don't match")
        print("   Recommendation: Manual approach")
        print("   Next step: Add descriptions to Excel manually")

if __name__ == "__main__":
    main()
