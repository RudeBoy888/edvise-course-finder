#!/usr/bin/env python3
"""
Scrape training.gov.au descriptions for ALL courses by name matching
Uses Selenium for JavaScript rendering
Only high-confidence matches (>85%) replace generated descriptions
"""

import openpyxl
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import json
from difflib import SequenceMatcher
from collections import defaultdict

def get_all_unique_courses():
    """Get all unique course names from Excel"""
    excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel not found")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws_courses = wb['Courses']

        courses = {}
        for row_idx in range(4, ws_courses.max_row + 1):
            expired = ws_courses.cell(row=row_idx, column=24).value
            if str(expired).strip().lower() != "no":
                continue

            course_name = ws_courses.cell(row=row_idx, column=4).value

            if course_name:
                course_name_clean = str(course_name).strip()
                if course_name_clean not in courses:
                    courses[course_name_clean] = []
                courses[course_name_clean].append(row_idx)

        return courses

    except Exception as e:
        print(f"❌ Error: {e}")
        return []

def search_course_on_nrt(course_name, driver):
    """Search for course on training.gov.au and extract description"""
    try:
        search_url = f"https://training.gov.au/search?searchText={course_name}&searchType=NRT"
        driver.get(search_url)
        time.sleep(1)

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        course_links = soup.find_all('a', href=True)

        matches = []
        for link in course_links:
            href = link.get('href', '')
            text = link.get_text(strip=True)

            if '/training/details/' in href and text:
                similarity = SequenceMatcher(None, course_name.lower(), text.lower()).ratio()
                if similarity > 0.75:
                    matches.append({
                        "name": text,
                        "url": href if href.startswith('http') else f"https://training.gov.au{href}",
                        "similarity": similarity
                    })

        if not matches:
            return None

        best_match = max(matches, key=lambda x: x['similarity'])

        # Only very confident matches
        if best_match['similarity'] < 0.85:
            return None

        # Fetch description
        driver.get(best_match['url'])
        time.sleep(0.5)

        soup = BeautifulSoup(driver.page_source, 'html.parser')

        description = None
        for heading in soup.find_all(['h2', 'h3', 'h4']):
            if 'Qualification description' in heading.get_text():
                next_elem = heading.find_next(['p', 'div'])
                if next_elem:
                    description = next_elem.get_text(strip=True)[:500]
                    break

        if not description:
            for p in soup.find_all('p'):
                text = p.get_text(strip=True)
                if len(text) > 100 and 'qualification' in text.lower():
                    description = text[:500]
                    break

        if description:
            return {
                "found_name": best_match['name'],
                "url": best_match['url'],
                "similarity": best_match['similarity'],
                "description": description
            }

        return None

    except Exception as e:
        return None

def main():
    print("🚀 FULL SCALE: Scraping training.gov.au for ALL courses\n")
    print("=" * 70)

    # Get all courses
    print("\n📚 Loading all unique course names...\n")
    courses = get_all_unique_courses()

    if not courses:
        print("❌ No courses found!")
        return

    total_courses = len(courses)
    print(f"✅ Found {total_courses} unique course names\n")

    # Setup Selenium
    print("🌐 Starting Selenium browser...\n")
    chrome_options = Options()
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')

    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
    except Exception as e:
        print(f"❌ Could not start Selenium: {e}")
        return

    # Search for courses
    results = {
        "found": [],
        "not_found": [],
        "stats": {}
    }

    start_time = time.time()

    for idx, (course_name, row_indices) in enumerate(courses.items(), 1):
        result = search_course_on_nrt(course_name, driver)

        if result:
            results["found"].append({
                "course_name": course_name,
                "rows": row_indices,
                **result
            })
            status = "✅"
        else:
            results["not_found"].append(course_name)
            status = "❌"

        if idx % 100 == 0:
            elapsed = time.time() - start_time
            rate = idx / elapsed
            remaining = (total_courses - idx) / rate if rate > 0 else 0
            print(f"[{idx}/{total_courses}] {status} Progress: {idx/total_courses*100:.1f}% | ETA: {remaining:.0f}s")

    driver.quit()

    # Results
    elapsed_total = time.time() - start_time
    print("\n" + "=" * 70)
    print(f"\n✅ SCRAPING COMPLETE ({elapsed_total:.0f}s)\n")
    print(f"Found: {len(results['found'])}/{total_courses} ({len(results['found'])/total_courses*100:.1f}%)")
    print(f"Not found: {len(results['not_found'])}/{total_courses}\n")

    if results["found"]:
        print("Sample matches:")
        for item in results["found"][:5]:
            print(f"  ✅ {item['course_name'][:60]} ({item['similarity']:.0%})")
            print(f"     → {item['description'][:80]}...")

    # Save results
    output_file = Path("/Users/rudybobek/edvise-course-finder/nrt_full_results.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"\n💾 Full results saved to: {output_file}")

    # Create merged descriptions CSV
    print("\n🔄 Creating merged descriptions file...\n")

    merged_file = Path("/Users/rudybobek/edvise-course-finder/courses_merged_descriptions.csv")

    # Map found courses by name
    found_map = {item['course_name']: item['description'] for item in results["found"]}

    with open(merged_file, 'w', encoding='utf-8') as f:
        f.write("Row,Course Name,Source,Description\n")

        excel_path = Path("/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52.xlsx")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws_courses = wb['Courses']

        for row_idx in range(4, ws_courses.max_row + 1):
            expired = ws_courses.cell(row=row_idx, column=24).value
            if str(expired).strip().lower() != "no":
                continue

            course_name = ws_courses.cell(row=row_idx, column=4).value

            if course_name:
                course_name_clean = str(course_name).strip()

                # Prefer real description from training.gov.au
                if course_name_clean in found_map:
                    description = found_map[course_name_clean]
                    source = "training.gov.au"
                else:
                    # Use generated description
                    generated_desc = ws_courses.cell(row=row_idx, column=25).value
                    description = str(generated_desc).strip() if generated_desc else ""
                    source = "generated"

                f.write(f'{row_idx},"{course_name_clean}","{source}","{description}"\n')

    print(f"   ✅ {merged_file}")

    print("\n" + "=" * 70)
    print(f"\n🎯 NEXT STEPS:\n")
    print(f"   1. Review: nrt_full_results.json")
    print(f"   2. Check: courses_merged_descriptions.csv")
    print(f"   3. Run: python3 import_merged_descriptions.py")
    print(f"   4. Run: python3 convert_excel_to_json.py")
    print(f"   5. Refresh app!\n")

if __name__ == "__main__":
    main()
