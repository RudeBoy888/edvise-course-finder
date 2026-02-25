#!/usr/bin/env python3
"""
Safe CRICOS Data Update Script
- Compares old vs new Excel data
- Generates change report
- Creates backups
- Updates data with validation
- Allows rollback if needed
"""

import json
import openpyxl
from pathlib import Path
from urllib.parse import urlparse
from datetime import datetime
import shutil
import hashlib

# Configuration
NEW_EXCEL_PATH = '/Users/rudybobek/Downloads/cricos-providers-courses-and-locations-as-at-2026-1-2-9-26-52 (1).xlsx'
OLD_JSON_PATH = 'public/courses_data.json'
BACKUP_DIR = 'data_backups'
REPORT_DIR = 'data_reports'
CURRENT_EXCEL = 'cricos_data_current.xlsx'

# Ensure backup and report directories exist
Path(BACKUP_DIR).mkdir(exist_ok=True)
Path(REPORT_DIR).mkdir(exist_ok=True)

def extract_domain(website_url):
    """Extract domain from full URL."""
    if not website_url or not str(website_url).strip():
        return None
    try:
        parsed = urlparse(str(website_url).strip())
        domain = parsed.netloc or parsed.path
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain if domain else None
    except:
        return None

def load_logo_mapping():
    """Load local logo mapping from JSON file."""
    try:
        with open('src/assets/logo-mapping.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def get_logo_url(domain, logo_mapping):
    """Get logo URL for a domain."""
    if not domain:
        return ""
    if domain in logo_mapping:
        return logo_mapping[domain]
    return ""

def load_institutions(ws, logo_mapping):
    """Load institutions from Institutions sheet."""
    institutions = {}
    for row_idx in range(4, ws.max_row + 1):
        provider_code = ws.cell(row=row_idx, column=1).value
        trading_name = ws.cell(row=row_idx, column=2).value
        institution_name = ws.cell(row=row_idx, column=3).value
        website = ws.cell(row=row_idx, column=6).value

        if not provider_code:
            continue

        name = trading_name.strip() if trading_name and trading_name.strip() else institution_name
        registered_name = institution_name.strip() if institution_name and institution_name.strip() else None

        domain = extract_domain(website)
        logo_url = get_logo_url(domain, logo_mapping)

        institution_data = {
            "name": name,
            "website": website if website else "",
            "domain": domain if domain else "",
            "logoUrl": logo_url if logo_url else ""
        }

        if registered_name and registered_name != name:
            institution_data["registeredName"] = registered_name

        institutions[str(provider_code).strip()] = institution_data

    return institutions

def load_locations(ws):
    """Load locations from Locations sheet."""
    locations = {}
    for row_idx in range(4, ws.max_row + 1):
        provider_code = ws.cell(row=row_idx, column=1).value
        location_name = ws.cell(row=row_idx, column=3).value
        address_line1 = ws.cell(row=row_idx, column=5).value
        address_line2 = ws.cell(row=row_idx, column=6).value
        city = ws.cell(row=row_idx, column=9).value
        state = ws.cell(row=row_idx, column=10).value
        postcode = ws.cell(row=row_idx, column=11).value

        if not provider_code or not location_name:
            continue

        address = ""
        if address_line1:
            address = str(address_line1).strip()
        if address_line2:
            address += " " + str(address_line2).strip()

        location_key = (str(provider_code).strip(), str(location_name).strip())
        locations[location_key] = {
            "address": address.strip() if address else "",
            "city": str(city).strip() if city else "",
            "state": str(state).strip() if state else "",
            "postcode": str(postcode).strip() if postcode else ""
        }

    return locations

def load_courses(ws):
    """Load courses from Courses sheet."""
    courses = []
    for row_idx in range(4, ws.max_row + 1):
        provider_code = ws.cell(row=row_idx, column=1).value
        course_code = ws.cell(row=row_idx, column=3).value
        course_name = ws.cell(row=row_idx, column=4).value
        duration_weeks = ws.cell(row=row_idx, column=20).value
        tuition_fee = ws.cell(row=row_idx, column=21).value
        non_tuition_fee = ws.cell(row=row_idx, column=22).value
        total_cost = ws.cell(row=row_idx, column=23).value
        field_of_education = ws.cell(row=row_idx, column=7).value
        course_level = ws.cell(row=row_idx, column=13).value
        foundation_studies = ws.cell(row=row_idx, column=14).value
        work_component = ws.cell(row=row_idx, column=15).value
        description = ws.cell(row=row_idx, column=25).value
        expired = ws.cell(row=row_idx, column=2).value

        if not provider_code or not course_code:
            continue
        if str(expired).strip().lower() == "yes":
            continue

        # Parse field of education
        if field_of_education:
            field_str = str(field_of_education).strip()
            if " - " in field_str:
                field_of_education = field_str.split(" - ", 1)[1]
            else:
                field_of_education = field_str

        # Convert fees to int
        try:
            tuition_fee = int(tuition_fee) if tuition_fee else None
        except:
            tuition_fee = None
        try:
            non_tuition_fee = int(non_tuition_fee) if non_tuition_fee else None
        except:
            non_tuition_fee = None
        try:
            total_cost = int(total_cost) if total_cost else None
        except:
            total_cost = None
        try:
            duration_weeks = int(duration_weeks) if duration_weeks else None
        except:
            duration_weeks = None

        # Convert booleans
        foundation_studies = str(foundation_studies).strip().lower() in ['yes', '1', 'true']
        work_component = str(work_component).strip().lower() in ['yes', '1', 'true']

        courses.append({
            "providerCode": str(provider_code).strip(),
            "courseCode": str(course_code).strip(),
            "courseName": str(course_name).strip() if course_name else "",
            "durationWeeks": duration_weeks,
            "tuitionFee": tuition_fee,
            "nonTuitionFee": non_tuition_fee,
            "totalCost": total_cost,
            "courseLevel": str(course_level).strip() if course_level else "",
            "fieldOfEducation": field_of_education,
            "isFoundationStudies": foundation_studies,
            "hasWorkComponent": work_component,
            "description": str(description).strip() if description else ""
        })

    return courses

def load_course_locations(ws):
    """Load course locations from Course Locations sheet."""
    course_locations = {}
    for row_idx in range(4, ws.max_row + 1):
        provider_code = ws.cell(row=row_idx, column=1).value
        course_code = ws.cell(row=row_idx, column=3).value
        location_name = ws.cell(row=row_idx, column=4).value
        city = ws.cell(row=row_idx, column=5).value
        state = ws.cell(row=row_idx, column=6).value

        if not provider_code or not course_code or not location_name:
            continue

        location_key = (str(provider_code).strip(), str(course_code).strip())
        location_item = {
            "locationName": str(location_name).strip(),
            "city": str(city).strip() if city else "",
            "state": str(state).strip() if state else ""
        }

        if location_key not in course_locations:
            course_locations[location_key] = []
        course_locations[location_key].append(location_item)

    return course_locations

def load_old_data():
    """Load old JSON data for comparison."""
    if not Path(OLD_JSON_PATH).exists():
        return None
    try:
        with open(OLD_JSON_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return None

def generate_change_report(old_data, new_data):
    """Generate detailed change report."""
    report = {
        "timestamp": datetime.now().isoformat(),
        "summary": {
            "old_institutions": len(old_data) if old_data else 0,
            "new_institutions": len(new_data),
            "old_total_courses": sum(len(i.get('courses', [])) for i in (old_data or [])),
            "new_total_courses": sum(len(i.get('courses', [])) for i in new_data),
        },
        "changes": {
            "added_institutions": [],
            "removed_institutions": [],
            "added_courses": [],
            "removed_courses": [],
            "modified_courses": [],
            "modified_institutions": []
        }
    }

    if not old_data:
        report["changes"]["added_institutions"] = [i['name'] for i in new_data]
        report["status"] = "initial_load"
        return report

    old_codes = {i['providerCode']: i for i in old_data}
    new_codes = {i['providerCode']: i for i in new_data}

    # Check for added/removed institutions
    for code in new_codes:
        if code not in old_codes:
            report["changes"]["added_institutions"].append(new_codes[code]['name'])

    for code in old_codes:
        if code not in new_codes:
            report["changes"]["removed_institutions"].append(old_codes[code]['name'])

    # Check for course changes
    for code in old_codes:
        if code in new_codes:
            old_inst = old_codes[code]
            new_inst = new_codes[code]

            old_courses = {c['courseCode']: c for c in old_inst.get('courses', [])}
            new_courses = {c['courseCode']: c for c in new_inst.get('courses', [])}

            # Check institution changes
            if old_inst.get('name') != new_inst.get('name'):
                report["changes"]["modified_institutions"].append({
                    "code": code,
                    "old_name": old_inst.get('name'),
                    "new_name": new_inst.get('name')
                })

            # Check for added/removed courses
            for course_code in new_courses:
                if course_code not in old_courses:
                    report["changes"]["added_courses"].append({
                        "institution": new_inst['name'],
                        "course": new_courses[course_code]['courseName'],
                        "code": course_code
                    })

            for course_code in old_courses:
                if course_code not in new_courses:
                    report["changes"]["removed_courses"].append({
                        "institution": old_inst['name'],
                        "course": old_courses[course_code]['courseName'],
                        "code": course_code
                    })

            # Check for modified courses
            for course_code in old_courses:
                if course_code in new_courses:
                    old_course = old_courses[course_code]
                    new_course = new_courses[course_code]
                    changes = {}

                    for key in ['courseName', 'tuitionFee', 'totalCost', 'durationWeeks']:
                        if old_course.get(key) != new_course.get(key):
                            changes[key] = {
                                "old": old_course.get(key),
                                "new": new_course.get(key)
                            }

                    if changes:
                        report["changes"]["modified_courses"].append({
                            "institution": new_inst['name'],
                            "course_code": course_code,
                            "course_name": new_course['courseName'],
                            "changes": changes
                        })

    report["status"] = "success"
    return report

def main():
    print("=" * 70)
    print("CRICOS DATA UPDATE - SAFE UPDATE PROCESS")
    print("=" * 70)

    # Step 1: Validate new Excel file
    print("\n[1/6] Validating new Excel file...")
    if not Path(NEW_EXCEL_PATH).exists():
        print(f"❌ ERROR: New Excel file not found: {NEW_EXCEL_PATH}")
        return False

    try:
        wb = openpyxl.load_workbook(NEW_EXCEL_PATH)
        print(f"✓ Excel file loaded successfully")
        print(f"  Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"❌ ERROR: Failed to load Excel: {e}")
        return False

    # Step 2: Load old data for comparison
    print("\n[2/6] Loading old data for comparison...")
    old_data = load_old_data()
    if old_data:
        print(f"✓ Old data loaded: {len(old_data)} institutions")
    else:
        print("⚠ No old data found (first time or missing file)")

    # Step 3: Parse new Excel data
    print("\n[3/6] Parsing new Excel data...")
    try:
        logo_mapping = load_logo_mapping()
        institutions = load_institutions(wb['Institutions'], logo_mapping)
        locations = load_locations(wb['Locations'])
        courses = load_courses(wb['Courses'])
        course_locations = load_course_locations(wb['Course Locations'])

        print(f"✓ Parsed: {len(institutions)} institutions, {len(courses)} courses")
    except Exception as e:
        print(f"❌ ERROR: Failed to parse Excel: {e}")
        return False

    # Step 4: Build new data structure
    print("\n[4/6] Building new data structure...")
    try:
        new_data = []
        for provider_code, inst_data in institutions.items():
            inst_courses = [c for c in courses if c['providerCode'] == provider_code]

            for course in inst_courses:
                location_key = (provider_code, course['courseCode'])
                course_locs = course_locations.get(location_key, [])

                course_locations_dict = {}
                for loc_item in course_locs:
                    state = loc_item.get('state', '')
                    if state not in course_locations_dict:
                        course_locations_dict[state] = []

                    location_key_tuple = (provider_code, loc_item['locationName'])
                    location_data = locations.get(location_key_tuple, {})

                    course_locations_dict[state].append({
                        "locationName": loc_item['locationName'],
                        "address": location_data.get('address', ''),
                        "city": loc_item.get('city', ''),
                        "state": state,
                        "postcode": location_data.get('postcode', '')
                    })

                course['locations'] = course_locations_dict

            if inst_courses:
                new_institution = {
                    "providerCode": provider_code,
                    **inst_data,
                    "allStates": list(set(
                        state for course in inst_courses
                        for state in course['locations'].keys()
                    )),
                    "courses": inst_courses
                }
                new_data.append(new_institution)

        print(f"✓ Built new data structure: {len(new_data)} institutions")
    except Exception as e:
        print(f"❌ ERROR: Failed to build data structure: {e}")
        return False

    # Step 5: Generate change report
    print("\n[5/6] Generating change report...")
    report = generate_change_report(old_data, new_data)

    print(f"✓ Report generated:")
    print(f"  - Institutions: {report['summary']['old_institutions']} → {report['summary']['new_institutions']}")
    print(f"  - Total courses: {report['summary']['old_total_courses']} → {report['summary']['new_total_courses']}")
    print(f"  - Added institutions: {len(report['changes']['added_institutions'])}")
    print(f"  - Removed institutions: {len(report['changes']['removed_institutions'])}")
    print(f"  - Added courses: {len(report['changes']['added_courses'])}")
    print(f"  - Removed courses: {len(report['changes']['removed_courses'])}")
    print(f"  - Modified courses: {len(report['changes']['modified_courses'])}")

    # Step 6: Backup old data and save new data
    print("\n[6/6] Saving new data...")
    try:
        # Create timestamped backup
        if Path(OLD_JSON_PATH).exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{BACKUP_DIR}/courses_data_{timestamp}.json"
            shutil.copy(OLD_JSON_PATH, backup_path)
            print(f"✓ Backup created: {backup_path}")

        # Save new JSON
        with open(OLD_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(new_data, f, ensure_ascii=False, indent=2)
        print(f"✓ New data saved: {OLD_JSON_PATH}")

        # Save report
        report_path = f"{REPORT_DIR}/update_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f"✓ Report saved: {report_path}")

        # Copy new Excel as reference
        if Path(CURRENT_EXCEL).exists():
            shutil.copy(CURRENT_EXCEL, f"{BACKUP_DIR}/{CURRENT_EXCEL.replace('.xlsx', '')}_old.xlsx")
        shutil.copy(NEW_EXCEL_PATH, CURRENT_EXCEL)
        print(f"✓ Excel reference saved: {CURRENT_EXCEL}")

    except Exception as e:
        print(f"❌ ERROR: Failed to save data: {e}")
        return False

    print("\n" + "=" * 70)
    print("✅ DATA UPDATE COMPLETED SUCCESSFULLY!")
    print("=" * 70)
    print(f"\nSummary:")
    print(f"  Old institutions: {report['summary']['old_institutions']}")
    print(f"  New institutions: {report['summary']['new_institutions']}")
    print(f"  Change: {report['summary']['new_institutions'] - report['summary']['old_institutions']:+d}")
    print(f"\n  Old total courses: {report['summary']['old_total_courses']}")
    print(f"  New total courses: {report['summary']['new_total_courses']}")
    print(f"  Change: {report['summary']['new_total_courses'] - report['summary']['old_total_courses']:+d}")

    return True

def run_html_report_generator():
    """Run the HTML report generator after data update."""
    import subprocess
    print("\n[7/7] Generating HTML report...")
    try:
        result = subprocess.run(['python3', 'generate_change_report_html.py'],
                              capture_output=True, text=True, timeout=30)
        if result.returncode == 0:
            print("✓ HTML report generated successfully")
            return True
        else:
            print(f"⚠ HTML report generation had issues: {result.stderr}")
            return False
    except Exception as e:
        print(f"⚠ Could not generate HTML report: {e}")
        return False

if __name__ == "__main__":
    success = main()
    if success:
        run_html_report_generator()
    exit(0 if success else 1)
